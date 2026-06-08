import json
import shutil
from pathlib import Path
from fastapi import APIRouter, UploadFile, File, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel
from urllib.parse import quote
from app.core.config import SQLITE_PATH, HOST, PORT, BASE_DIR
import openpyxl

try:
    import mammoth
except ImportError:
    mammoth = None

router = APIRouter(prefix="/api/system", tags=["system"])
SETTINGS_FILE = BASE_DIR / "data" / "system_settings.json"

# ==============================================================================
# 新增：管理员账号配置文件逻辑
# ==============================================================================
ADMIN_FILE = BASE_DIR / "admin_account.json"

def get_admin_credentials():
    # 如果根目录没有这个文件，就自动创建并写入初始账号密码
    if not ADMIN_FILE.exists():
        default_creds = {"username": "admin", "password": "liqi030530"}
        ADMIN_FILE.write_text(json.dumps(default_creds, ensure_ascii=False, indent=2), encoding="utf-8")
        return default_creds
    try:
        return json.loads(ADMIN_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"username": "admin", "password": "liqi030530"}

class LoginPayload(BaseModel):
    username: str
    password: str

@router.post("/login")
def login_admin(payload: LoginPayload):
    creds = get_admin_credentials()
    if payload.username == creds.get("username") and payload.password == creds.get("password"):
        return {"success": True, "message": "管理员验证成功"}
    return {"success": False, "message": "账号或密码错误，请检查！"}

# ==============================================================================

class SettingsPayload(BaseModel):
    sourceDir: str = "data/source"
    syncMode: str = "manual"
    cron: str = "0 0/30 * * * *"

def load_settings():
    if SETTINGS_FILE.exists():
        try:
            return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except Exception: pass
    return {"sourceDir": "data/source", "syncMode": "manual", "cron": "0 0/30 * * * *"}

def save_settings_data(data: dict):
    SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    SETTINGS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

@router.get("/status")
def system_status():
    return {"host": HOST, "port": PORT, "db_path": str(SQLITE_PATH)}

@router.get("/settings")
def get_settings():
    return load_settings()

@router.post("/settings")
def save_settings(payload: SettingsPayload):
    save_settings_data(payload.model_dump())
    return {"success": True, "message": "设置已保存"}

@router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    filename = file.filename
    if not filename: 
        return {"success": False, "message": "无法获取文件名，上传失败"}

    fn_lower = filename.lower()
    
    if fn_lower.endswith('.xlsx'):
        settings = load_settings()
        source_dir_str = settings.get("sourceDir", "data/source")
        source_path = Path(source_dir_str)
        dest_dir = BASE_DIR / source_path if not source_path.is_absolute() else source_path
        msg_suffix = "已存入台账待同步目录"
        
    elif fn_lower.endswith('.pdf'):
        dest_dir = BASE_DIR / "data" / "standards"
        msg_suffix = "已存入企业标准原件库"
    else:
        return {"success": False, "message": "不支持的文件类型"}

    dest_dir.mkdir(parents=True, exist_ok=True)
    file_path = dest_dir / filename

    try:
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"success": True, "message": f"文件 {filename} 上传成功，{msg_suffix}"}
    except Exception as e:
        return {"success": False, "message": f"文件保存失败: {str(e)}"}


@router.get("/docs")
def list_system_docs():
    docs_dir = BASE_DIR / "data" / "system"
    if not docs_dir.exists():
        docs_dir.mkdir(parents=True, exist_ok=True)
        
    allowed_exts = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}
    items = []
    
    for p in docs_dir.rglob("*"):
        if p.is_file() and p.suffix.lower() in allowed_exts:
            rel_path = p.relative_to(docs_dir).as_posix()
            items.append({
                "filename": p.name,
                "rel_path": rel_path,
                "ext": p.suffix.lower()
            })
            
    items.sort(key=lambda x: x["filename"])
    return {"success": True, "items": items}

@router.get("/docs/file")
def get_system_doc_file(path: str = Query(...), download: bool = Query(False)):
    docs_dir = BASE_DIR / "data" / "system"
    file_path = (docs_dir / path).resolve()
    
    if not str(file_path).startswith(str(docs_dir.resolve())):
        return {"success": False, "message": "非法的文件路径"}
        
    if file_path.exists() and file_path.is_file():
        headers = {}
        if download:
            encoded_name = quote(file_path.name)
            headers["Content-Disposition"] = f"attachment; filename*=utf-8''{encoded_name}"
        return FileResponse(str(file_path), headers=headers)
        
    return {"success": False, "message": "请求的文件不存在或已被移除"}

@router.get("/docs/preview-html")
def get_office_preview(path: str = Query(...)):
    docs_dir = BASE_DIR / "data" / "system"
    file_path = (docs_dir / path).resolve()
    
    if not str(file_path).startswith(str(docs_dir.resolve())):
        return {"success": False, "message": "非法的文件路径"}
    
    if not file_path.exists() or not file_path.is_file():
        return {"success": False, "message": "文件不存在"}
        
    ext = file_path.suffix.lower()
    
    if ext == '.docx':
        if not mammoth:
            return {"success": False, "message": "系统未安装 mammoth 引擎，无法解析 Word，请联系管理员"}
        try:
            with open(file_path, "rb") as docx_file:
                result = mammoth.convert_to_html(docx_file)
                html_content = f"<div style='padding: 20px; font-family: sans-serif; line-height: 1.6; max-width: 900px; margin: 0 auto; background: #fff; box-shadow: 0 1px 3px rgba(0,0,0,0.1);'>{result.value}</div>"
                return {"success": True, "html": html_content}
        except Exception as e:
            return {"success": False, "message": f"Word文档解析失败: {str(e)}"}
            
    elif ext == '.xlsx':
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            html_parts = ["<div style='padding: 16px; font-family: sans-serif; background: #fff;'>"]
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                html_parts.append(f"<h3 style='background:#f1f5f9; padding:10px 16px; margin: 20px 0 10px 0; border-left: 4px solid #3b82f6; color: #1e293b;'>工作表：{sheet_name}</h3>")
                html_parts.append("<div style='overflow-x:auto; box-shadow: 0 0 0 1px #e2e8f0; border-radius: 6px;'><table style='border-collapse: collapse; width: 100%; text-align: left; font-size: 13px;'>")
                
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if idx > 200:
                        html_parts.append("<tr><td colspan='20' style='padding:12px; color:#64748b; text-align:center; background:#f8fafc;'>... 篇幅限制，仅展示前 200 行用于快速预览 ...</td></tr>")
                        break
                        
                    html_parts.append("<tr>")
                    for cell_idx, cell in enumerate(row):
                        val = "" if cell is None else str(cell)
                        bg_color = "#f8fafc" if idx == 0 else "#ffffff"
                        font_weight = "bold" if idx == 0 else "normal"
                        html_parts.append(f"<td style='border: 1px solid #e2e8f0; padding: 8px 12px; white-space: nowrap; background: {bg_color}; font-weight: {font_weight};'>{val}</td>")
                    html_parts.append("</tr>")
                    
                html_parts.append("</table></div>")
            
            html_parts.append("</div>")
            wb.close()
            return {"success": True, "html": "".join(html_parts)}
        except Exception as e:
             return {"success": False, "message": f"Excel表格解析失败: {str(e)}"}
             
    return {"success": False, "message": "暂不支持该后缀文件的直接渲染"}