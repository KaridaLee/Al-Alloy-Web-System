import hashlib
import re
from datetime import datetime
from openpyxl import load_workbook

ELEMENTS_ORDER = [
    "Al", "Si", "Cu", "Mg", "Mn", "Ti", "Fe", "Zn", "Ni", "Pb", "Sn",
    "Sr", "Zr", "Cr", "Ca", "Sb", "Cd", "As", "B", "Be", "Bi", "Co",
    "Ga", "Hg", "Li", "Mo", "Na", "P", "V"
]
ELEMENTS = set(ELEMENTS_ORDER)

BASE_FIELDS_CANONICAL = [
    "序号", "炉号", "牌号", "批次号", "班组_班长", "针孔度判定", "检测时间时间"
]

TAIL_FIELDS_TEMPLATE = [
    "判定", "判定人", "发货", "备注", "烂泥指数", "低于内控", "高于内控", "低于内控_高于内控"
]

PREFERRED_KEY_GROUPS = [
    ["炉号", "批次号"],
    ["炉号", "牌号", "批次号"],
    ["序号"],
    ["炉号", "检测时间时间"],
]


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep="T", timespec="seconds")
    return str(value).strip().replace("\n", "").replace("\r", "")


def normalize_column_name(name: str) -> str:
    if not name:
        return ""
    name = str(name).strip()
    name = name.replace("/", "_")
    name = name.replace("\n", "")
    name = name.replace("\r", "")
    name = name.replace(" ", "")
    return name


def fill_merged_value(sheet, row, col):
    value = sheet.cell(row, col).value
    if value is not None:
        return value

    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return None


def is_non_empty(value):
    return normalize_text(value) != ""


def normalize_element_header(value: str):
    """
    将元素类表头统一规范化：
    兼容中英文括号、特殊符号以及前后缀干扰。
    例如：
    - Cu
    - Cu（1.5-3.5）
    - Si(9.6-12)
    - P(0.0035-0.0070)
    统一映射为纯元素符号
    """
    text = normalize_text(value)
    if not text:
        return ""

    # 去掉所有空白，便于统一识别
    compact = re.sub(r"\s+", "", text)

    # 先直接检查纯元素
    if compact in ELEMENTS:
        return compact

    # 使用高兼容性的穿透正则：提取首个元素符号，忽略其后的任何非字母字符（如中英文括号、数字、范围符号）
    match = re.match(r'^[^a-zA-Z]*([A-Z][a-z]?)(?:[^a-zA-Z].*)?$', compact)
    if match:
        elem = match.group(1)
        if elem in ELEMENTS:
            return elem

    return ""


def detect_header_rows(sheet):
    """
    你的模板基本固定：
    第1行标题
    第2-3行表头
    第4行开始数据
    """
    if sheet.max_row >= 4:
        return 2, 3, 4
    return 1, 1, 2


def detect_effective_max_col(sheet, header_top_row, header_bottom_row, max_scan_cols=120, empty_limit=12):
    max_col = min(sheet.max_column, max_scan_cols)

    last_non_empty_col = 0
    consecutive_empty = 0

    probe_rows = [header_top_row, header_bottom_row]
    if header_bottom_row + 1 <= sheet.max_row:
        probe_rows.append(header_bottom_row + 1)

    for c in range(1, max_col + 1):
        has_value = False
        for r in probe_rows:
            if is_non_empty(sheet.cell(r, c).value):
                has_value = True
                break

        if has_value:
            last_non_empty_col = c
            consecutive_empty = 0
        else:
            consecutive_empty += 1
            if consecutive_empty >= empty_limit and last_non_empty_col > 0:
                break

    if last_non_empty_col == 0:
        last_non_empty_col = min(max_col, 40)

    return last_non_empty_col


def canonical_base_field(top_val, bottom_val):
    top_val = normalize_column_name(top_val)
    bottom_val = normalize_column_name(bottom_val)

    combined = f"{top_val}_{bottom_val}".strip("_")

    if top_val == "序号" or bottom_val == "序号":
        return "序号"
    if top_val == "炉号" or bottom_val == "炉号":
        return "炉号"
    if top_val == "牌号" or bottom_val == "牌号":
        return "牌号"
    if top_val == "批次号" or bottom_val == "批次号":
        return "批次号"
    if top_val in ("班组_班长", "班组/班长") or bottom_val in ("班组_班长", "班组/班长"):
        return "班组_班长"
    if top_val == "针孔度判定" or bottom_val == "针孔度判定":
        return "针孔度判定"

    if top_val == "检测时间" and bottom_val == "时间":
        return "检测时间时间"
    if combined == "检测时间_时间":
        return "检测时间时间"
    if top_val == "检测时间时间" or bottom_val == "检测时间时间":
        return "检测时间时间"
    if top_val == "检测时间" and not bottom_val:
        return "检测时间时间"

    return ""


def build_headers_fixed(sheet, header_top_row, header_bottom_row, effective_max_col):
    columns = []
    col_map = []

    raw_info = []
    for c in range(1, effective_max_col + 1):
        top_val = normalize_text(fill_merged_value(sheet, header_top_row, c))
        bottom_val = normalize_text(fill_merged_value(sheet, header_bottom_row, c))

        raw_info.append({
            "col": c,
            "top": top_val,
            "bottom": bottom_val
        })

    # 找元素区
    element_positions = []
    for info in raw_info:
        normalized_elem = normalize_element_header(info["bottom"])
        if normalized_elem:
            element_positions.append(info["col"])

    first_element_col = min(element_positions) if element_positions else None
    last_element_col = max(element_positions) if element_positions else None

    seen = {}

    def add_col(col_index, name):
        if not name:
            return
        final_name = name
        if final_name in seen:
            seen[final_name] += 1
            final_name = f"{final_name}_{seen[final_name]}"
        else:
            seen[final_name] = 1

        columns.append(final_name)
        col_map.append((col_index, final_name))

    # 前置基础字段
    for info in raw_info:
        c = info["col"]
        if first_element_col and c >= first_element_col:
            break

        field_name = canonical_base_field(info["top"], info["bottom"])
        if field_name:
            add_col(c, field_name)

    # 元素字段
    if first_element_col and last_element_col:
        for info in raw_info:
            c = info["col"]
            if c < first_element_col or c > last_element_col:
                continue

            elem = normalize_element_header(info["bottom"])
            if elem:
                add_col(c, elem)

    # 元素区后的尾部字段
    if last_element_col:
        tail_start = last_element_col + 1
        tail_fields_idx = 0

        for c in range(tail_start, effective_max_col + 1):
            if tail_fields_idx >= len(TAIL_FIELDS_TEMPLATE):
                break

            top_val = normalize_text(fill_merged_value(sheet, header_top_row, c))
            bottom_val = normalize_text(fill_merged_value(sheet, header_bottom_row, c))

            if not normalize_column_name(top_val) and not normalize_column_name(bottom_val):
                continue

            add_col(c, TAIL_FIELDS_TEMPLATE[tail_fields_idx])
            tail_fields_idx += 1
    else:
        # 兜底
        for info in raw_info:
            candidate = normalize_column_name(info["bottom"]) or normalize_column_name(info["top"])
            if candidate and "台账" not in candidate and not candidate.startswith("化学成分"):
                add_col(info["col"], candidate)

    return columns, col_map


def choose_key_columns(columns):
    col_set = set(columns)

    for group in PREFERRED_KEY_GROUPS:
        if all(g in col_set for g in group):
            return group

    if "序号" in col_set:
        return ["序号"]

    if columns:
        return [columns[0]]

    return []


def find_furnace_col_name(columns):
    return "炉号" if "炉号" in columns else None


def build_row_key(row_dict, key_columns, fallback_index):
    parts = [str(row_dict.get(col, "")).strip() for col in key_columns]
    joined = "|".join(parts).strip("|").strip()
    if joined:
        return joined
    return f"__fallback__{fallback_index}"


def build_row_hash(row_dict):
    items = []
    for k in sorted(row_dict.keys()):
        if k == "__row_key":
            continue
        items.append(f"{k}={row_dict.get(k, '')}")
    raw = "||".join(items)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def is_data_row_valid(row_dict, furnace_col_name=None):
    if furnace_col_name and furnace_col_name in row_dict:
        if str(row_dict.get(furnace_col_name, "")).strip() != "":
            return True

    return any(str(v).strip() != "" for v in row_dict.values())


def read_effective_rows(sheet, data_start_row, col_map, key_columns, furnace_col_name=None, empty_limit=3):
    rows = []
    fallback_index = 1
    consecutive_empty = 0

    for r in range(data_start_row, sheet.max_row + 1):
        row_dict = {}
        for col_idx, col_name in col_map:
            row_dict[col_name] = normalize_text(sheet.cell(r, col_idx).value)

        if not is_data_row_valid(row_dict, furnace_col_name=furnace_col_name):
            consecutive_empty += 1
            if consecutive_empty >= empty_limit:
                break
            continue

        consecutive_empty = 0

        row_key = build_row_key(row_dict, key_columns, fallback_index)
        fallback_index += 1

        row_dict["__row_key"] = row_key
        row_hash = build_row_hash(row_dict)

        rows.append({
            "row_key": row_key,
            "row_hash": row_hash,
            "data": row_dict
        })

    return rows


def parse_workbook(file_path: str):
    wb = load_workbook(file_path, data_only=True)
    workbook_result = []

    for ws in wb.worksheets:
        print(f"[PARSE] 开始分析工作表: {ws.title}")
        print(f"[PARSE] 原始 max_row={ws.max_row}, max_column={ws.max_column}")

        header_top_row, header_bottom_row, data_start_row = detect_header_rows(ws)
        print(f"[PARSE] header_top_row={header_top_row}, header_bottom_row={header_bottom_row}, data_start_row={data_start_row}")

        effective_max_col = detect_effective_max_col(
            ws,
            header_top_row,
            header_bottom_row,
            max_scan_cols=120,
            empty_limit=12
        )
        print(f"[PARSE] 动态识别有效列数={effective_max_col}")

        columns, col_map = build_headers_fixed(ws, header_top_row, header_bottom_row, effective_max_col)
        print(f"[PARSE] 识别字段数={len(columns)}")
        print(f"[PARSE] 字段预览={columns[:40]}")

        key_columns = choose_key_columns(columns)
        furnace_col_name = find_furnace_col_name(columns)

        print(f"[PARSE] 选定 key_columns={key_columns}")
        print(f"[PARSE] 识别炉号列={furnace_col_name}")

        rows = read_effective_rows(
            ws,
            data_start_row,
            col_map,
            key_columns,
            furnace_col_name=furnace_col_name,
            empty_limit=3
        )

        print(f"[PARSE] 工作表 {ws.title} 最终有效数据行数={len(rows)}")

        workbook_result.append({
            "sheet_name": ws.title,
            "table_name": f"sheet_{ws.title}".replace(" ", "_").replace("-", "_"),
            "header_top_row": header_top_row,
            "header_bottom_row": header_bottom_row,
            "data_start_row": data_start_row,
            "effective_max_col": effective_max_col,
            "columns": columns,
            "key_columns": key_columns,
            "rows": rows
        })

    return workbook_result